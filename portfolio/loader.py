"""매매 이력(구매 이력) 파일 로딩.

휴대폰 메모/엑셀/구글시트에서 내보낸 파일을 그대로 받도록 컬럼명을 유연하게 처리한다.
지원 포맷: .csv, .tsv, .json, .xlsx(openpyxl 있을 때)
한글 헤더도 인식: 날짜/종목/구분/수량/단가/수수료/계좌/메모
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import json
import re
from pathlib import Path

from .models import Transaction

# 컬럼 별칭 -> 표준 필드
ALIASES: dict[str, str] = {}


def _alias(field: str, *names: str) -> None:
    for n in names:
        ALIASES[_norm(n)] = field


def _norm(s: str) -> str:
    return re.sub(r"[\s_\-/().]", "", str(s or "")).lower()


_alias("date", "date", "날짜", "거래일", "매매일자", "체결일", "일자", "tradedate")
_alias("ticker", "ticker", "symbol", "종목", "종목코드", "티커", "code", "종목명")
_alias("side", "side", "type", "구분", "매매구분", "거래구분", "action", "매매")
_alias("quantity", "quantity", "qty", "shares", "수량", "주수", "주식수", "체결수량")
_alias("price", "price", "단가", "가격", "매입가", "체결단가", "평단", "unitprice")
_alias("fee", "fee", "commission", "수수료", "제비용", "세금", "tax")
_alias("account", "account", "계좌", "증권사", "broker", "계좌명")
_alias("note", "note", "memo", "메모", "비고", "comment")

BUY_WORDS = {"buy", "b", "매수", "구매", "buy1", "매입", "+"}
SELL_WORDS = {"sell", "s", "매도", "판매", "-"}


class LoaderError(RuntimeError):
    pass


def load_transactions(path: Path | str) -> list[Transaction]:
    path = Path(path)
    if not path.exists():
        raise LoaderError(f"거래내역 파일이 없습니다: {path}")
    suffix = path.suffix.lower()
    if suffix == ".json":
        rows = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(rows, dict):
            rows = rows.get("transactions") or []
    elif suffix in (".xlsx", ".xlsm"):
        rows = _read_xlsx(path)
    else:
        rows = _read_delimited(path)

    txs: list[Transaction] = []
    for i, row in enumerate(rows, start=2):
        tx = _row_to_tx(row, path.name, i)
        if tx:
            txs.append(tx)
    if not txs:
        raise LoaderError(f"읽을 수 있는 거래가 없습니다: {path}")
    txs.sort(key=lambda t: (t.date, t.ticker))
    return txs


def _read_delimited(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8-sig")
    # 구분자 자동 감지 (엑셀에서 내보낸 탭 구분 파일 대응)
    sample = text[:2048]
    delim = "\t" if sample.count("\t") > sample.count(",") else ","
    return list(csv.DictReader(io.StringIO(text), delimiter=delim))


def _read_xlsx(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise LoaderError("xlsx 를 읽으려면 `pip install openpyxl` 이 필요합니다.") from e
    ws = load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h or "") for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:]]


def _row_to_tx(row: dict, source: str, lineno: int) -> Transaction | None:
    data: dict[str, object] = {}
    for k, v in row.items():
        field = ALIASES.get(_norm(k))
        if field and (v is not None) and str(v).strip() != "":
            data.setdefault(field, v)
    if not data.get("ticker") or data.get("quantity") in (None, ""):
        return None

    side_raw = _norm(data.get("side", "buy"))
    if side_raw in SELL_WORDS:
        side = "SELL"
    elif side_raw in BUY_WORDS or not side_raw:
        side = "BUY"
    else:
        raise LoaderError(f"{source}:{lineno} 알 수 없는 매매구분 '{data.get('side')}'")

    try:
        qty = abs(_num(data["quantity"]))
        price = _num(data.get("price", 0))
    except ValueError as e:
        raise LoaderError(f"{source}:{lineno} 숫자 변환 실패: {e}") from e

    return Transaction(
        date=_date(data.get("date"), source, lineno),
        ticker=str(data["ticker"]).strip(),
        side=side,
        quantity=qty,
        price=price,
        fee=_num(data.get("fee", 0)) if data.get("fee") else 0.0,
        account=str(data.get("account", "기본")).strip() or "기본",
        note=str(data.get("note", "")).strip(),
    )


def _num(v) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("₩", "").replace("$", "").replace("원", "")
    if not s:
        return 0.0
    return float(s)


def _date(v, source: str, lineno: int) -> dt.date:
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v or "").strip()
    if not s:
        return dt.date.today()
    s = s.split(" ")[0].split("T")[0]
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%d-%m-%Y", "%m/%d/%Y", "%Y%m%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise LoaderError(f"{source}:{lineno} 날짜 형식을 못 읽었습니다: '{v}'")


def append_transactions(target: Path | str, txs: list[Transaction]) -> int:
    """가져온 거래를 표준 CSV 에 덧붙인다 (중복은 건너뜀)."""
    target = Path(target)
    existing: set[tuple] = set()
    if target.exists():
        for t in load_transactions(target):
            existing.add((t.date, t.ticker, t.side, t.quantity, t.price, t.account))
    new_rows = [
        t for t in txs
        if (t.date, t.ticker, t.side, t.quantity, t.price, t.account) not in existing
    ]
    if not new_rows:
        return 0
    write_header = not target.exists()
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("a", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        if write_header:
            w.writerow(["date", "ticker", "side", "quantity", "price", "fee", "account", "note"])
        for t in new_rows:
            w.writerow([t.date.isoformat(), t.ticker, t.side, _fmt(t.quantity),
                        _fmt(t.price), _fmt(t.fee), t.account, t.note])
    return len(new_rows)


def _fmt(v: float) -> str:
    return str(int(v)) if float(v).is_integer() else f"{v:g}"
