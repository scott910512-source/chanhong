# -*- coding: utf-8 -*-
"""검산용: 예시 숫자를 채운 사본을 만든다.
GOOGLEFINANCE 는 구글시트 전용이라 검산 엔진이 모른다. 그 세 칸(M/N/O)만
빼고 나머지 수식이 맞는지 본다. K(현재가)는 종목코드를 비워 '수동 현재가'
경로로 흐르게 해서, 실제 쓰는 길과 같은 식을 그대로 태운다."""
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
wb = load_workbook(ROOT / '찬홍팍_자산배분_구글시트.xlsx')
ws = wb['자산배분']
for r in range(6, 31):
    for c in 'MNO':
        if ws[f'{c}{r}'].value: ws[f'{c}{r}'] = None

PRICE = {'ACE 미국나스닥100': 21500, 'TIGER  미국S&P500': 19800,
         'SOL 미국배당다우존스': 11200, 'TIGER  미국배당다우존스': 11450,
         'KODEX 미국나스닥100': 18300, 'KODEX 미국S&P500': 17600,
         '1Q미국S&P500미국채혼합50액티브': 12100, 'KODEX 200TR': 13900,
         'TIGER 일본니케이225': 22400, 'KODEX 차이나CSI300': 10800,
         'KODEX 인도Nifty50': 14700, '매그니피센트7 MAGS': 78500,
         'Schwab SCHD': 38200, '버크셔 해서웨이B BRK.B': 720000,
         '금 99.99K': 168000, '비트코인': 152000000, '이더리움': 5400000, '현금': 1}
for r in range(6, 31):
    n = ws[f'I{r}'].value
    if n in PRICE: ws[f'X{r}'] = PRICE[n]

# 계좌 세 곳에만 실제로 채워 넣는다
HOLD = {2: (20000, 120), 3: (18500, 90), 4: (10500, 200), 5: (13000, 150),
        6: (21000, 40), 7: (10200, 80), 8: (13800, 60),
        16: (19800, 55), 17: (18000, 40), 18: (11000, 70), 19: (13500, 30),
        51: (150000, 12), 52: (140000000, 0.05), 53: (5000000, 0.4), 54: (1, 8000000)}
for r, (buy, qty) in HOLD.items():
    ws[f'D{r}'] = buy; ws[f'E{r}'] = qty
wb.save(ROOT / '검산용.xlsx'); print('검산용.xlsx 준비')
