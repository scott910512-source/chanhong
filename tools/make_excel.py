# -*- coding: utf-8 -*-
"""찬홍팍 주식관리 엑셀을 만든다.

앱(web/)과 같은 계산 규칙을 엑셀 수식으로 옮긴 관리용 파일.
  python3 tools/make_excel.py            # 저장소 루트에 찬홍팍_주식관리.xlsx 생성

수식 검산은 tools/check_excel.py 로 한다 (실제로 계산해서 값을 대조한다).
이 상자에서는 리브레오피스가 소켓이 막혀 안 돌아가서 recalc 대신 그걸 쓴다.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

F = 'Arial'
INK   = Font(name=F, size=10)                                  # 수식(검정)
IN    = Font(name=F, size=10, color='0000FF')                  # 직접 입력(파랑)
LINK  = Font(name=F, size=10, color='008000')                  # 다른 시트 참조(초록)
H1    = Font(name=F, size=14, bold=True)
H2    = Font(name=F, size=11, bold=True)
TH    = Font(name=F, size=10, bold=True, color='FFFFFF')
NOTE  = Font(name=F, size=9, color='666666')
HFILL = PatternFill('solid', fgColor='2563EB')
IFILL = PatternFill('solid', fgColor='FFF7CC')                 # 입력 칸
SFILL = PatternFill('solid', fgColor='EEF2F7')
THIN  = Side(style='thin', color='D0D5DD')
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY, PRICE, QTY, PCT = '#,##0', '#,##0.####', '#,##0.####', '0.0%'

TXN_TOP, TXN_END = 4, 253      # 거래입력 데이터 행
AST_TOP, AST_END = 4, 63       # 종목정보 데이터 행
POS_TOP = 6                    # 보유현황 데이터 시작
POS_END = POS_TOP + (AST_END - AST_TOP)
TGT_TOP, TGT_END = 4, 23       # 국가별/섹터별 목표 행

wb = Workbook()

def head(ws, title, sub, widths):
    ws['A1'] = title; ws['A1'].font = H1
    ws['A2'] = sub;   ws['A2'].font = NOTE
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.sheet_view.showGridLines = False

def hrow(ws, row, labels, start=1):
    for i, t in enumerate(labels):
        c = ws.cell(row=row, column=start + i, value=t)
        c.font = TH; c.fill = HFILL; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 30

# ══════════════════════════════════════════ 1. 사용법
ws = wb.active; ws.title = '사용법'
head(ws, '찬홍팍 주식관리 엑셀', '', {'A': 22, 'B': 96})
ws['A2'] = '앱(chanhong)과 같은 계산 규칙을 엑셀 수식으로 옮긴 파일입니다.'
GUIDE = [
    ('', ''),
    ('■ 손대는 곳', '노란 칸 + 파란 글씨. 이 셋만 채우면 나머지는 전부 자동으로 계산됩니다.'),
    ('  1. 종목정보', '가진 종목을 등록합니다. 티커 / 종목명 / 국가 / 통화 / 섹터 / 현재가'),
    ('  2. 거래입력', '산 날, 판 날을 한 줄씩 적습니다. 반드시 날짜 순으로 적으세요.'),
    ('  3. 설정', '환율과 허용오차. 국가별·섹터별 시트의 목표 비중도 여기서 정합니다.'),
    ('', ''),
    ('■ 손대지 마세요', '검은 글씨 = 수식입니다. 지우면 계산이 깨집니다.'),
    ('', '초록 글씨 = 다른 시트에서 끌어온 값입니다.'),
    ('', ''),
    ('■ 자동으로 나오는 것', ''),
    ('  보유현황', '종목별 보유수량 / 평단 / 평가금액 / 손익 / 수익률 / 비중'),
    ('  국가별', '국가별 비중과 목표 대비 판정 (비중 많음 / 적정 / 비중 적음)'),
    ('  섹터별', '섹터별 비중과 목표 대비 판정'),
    ('', ''),
    ('■ 계산 규칙 (앱과 같음)', ''),
    ('  평단', '이동평균. 살 때 수수료를 매입원가에 포함합니다.'),
    ('', '팔아도 평단은 그대로 두고, 판 몫의 손익만 실현손익으로 뺍니다.'),
    ('  시세가 없으면', '평가금액을 0 으로 두지 않고 산 값(매입가)으로 칩니다.'),
    ('', '0 으로 두면 방금 넣은 종목이 총자산에서 통째로 사라집니다.'),
    ('  비중', '모든 금액을 기준통화(원)로 바꾼 뒤 계산합니다.'),
    ('  판정', '목표 ± 허용오차 를 벗어나면 많음 / 적음으로 표시합니다.'),
    ('', ''),
    ('■ 앱에서 옮겨오기', '앱 → 설정 → 엑셀로 내려받기 → [거래내역] 시트를 통째로 복사해서'),
    ('', '이 파일 거래입력 시트 A4 에 붙여넣으면 됩니다. 열 순서가 같습니다.'),
    ('', ''),
    ('■ 예시 데이터', '지금 들어있는 숫자는 예시입니다. 거래입력·종목정보의 내용을 지우고'),
    ('', '형님 것으로 채우세요. 수식이 있는 칸은 지우지 마세요.'),
    ('', ''),
    ('■ 넣을 수 있는 양', '거래 250건 / 종목 60개. 모자라면 마지막 행을 복사해 아래로 늘리세요.'),
]
r = 4
for a, b in GUIDE:
    ws.cell(row=r, column=1, value=a).font = H2 if a.startswith('■') else INK
    c = ws.cell(row=r, column=2, value=b); c.font = INK
    c.alignment = Alignment(vertical='top')
    r += 1
ws['A' + str(r + 1)] = '색깔 규칙'
ws['A' + str(r + 1)].font = H2
for i, (t, f) in enumerate([('직접 입력하는 칸', IN), ('수식 (건드리지 마세요)', INK),
                            ('다른 시트에서 온 값', LINK)]):
    c = ws.cell(row=r + 2 + i, column=1, value='■■■'); c.font = f
    ws.cell(row=r + 2 + i, column=2, value=t).font = INK
ws.cell(row=r + 2, column=1).fill = IFILL

# ══════════════════════════════════════════ 2. 설정
ws = wb.create_sheet('설정')
head(ws, '설정', '환율과 허용오차. 노란 칸만 고치세요.',
     {'A': 16, 'B': 14, 'C': 46})
ws['A4'] = '기준통화'; ws['A4'].font = H2
ws['B4'] = 'KRW'; ws['B4'].font = IN; ws['B4'].fill = IFILL; ws['B4'].border = BOX
ws['C4'] = '모든 금액을 이 통화로 바꿔서 비중을 냅니다.'; ws['C4'].font = NOTE

ws['A6'] = '환율'; ws['A6'].font = H2
ws['C6'] = '1 외화 = 몇 원인지 적으세요. 앱 → 설정 → 환율 에서 볼 수 있습니다.'
ws['C6'].font = NOTE
hrow(ws, 7, ['통화', '원화 환산', '메모'])
FX = [('KRW', 1, '기준통화'), ('USD', 1340, '미국 달러'), ('VND', 0.0525, '베트남 동'),
      ('JPY', 9.1, '일본 엔'), ('EUR', 1450, '유로'), ('HKD', 172, '홍콩 달러'),
      ('CNY', 186, '중국 위안'), ('GBP', 1700, '영국 파운드')]
for i, (cur, rate, memo) in enumerate(FX):
    rr = 8 + i
    for col, val, fnt in ((1, cur, IN), (2, rate, IN), (3, memo, NOTE)):
        c = ws.cell(row=rr, column=col, value=val); c.font = fnt; c.border = BOX
        if col <= 2: c.fill = IFILL
    ws.cell(row=rr, column=2).number_format = '#,##0.######'
FX_A, FX_B = 8, 8 + len(FX) - 1

ws['A20'] = '허용오차'; ws['A20'].font = H2
ws['C20'] = '목표에서 이만큼까지는 "적정" 으로 봅니다. 5%p 면 0.05 를 넣으세요.'
ws['C20'].font = NOTE
for i, (lbl, val) in enumerate([('국가별', 0.05), ('섹터별', 0.05)]):
    rr = 21 + i
    ws.cell(row=rr, column=1, value=lbl).font = INK
    c = ws.cell(row=rr, column=2, value=val)
    c.font = IN; c.fill = IFILL; c.border = BOX; c.number_format = PCT
TOL_COUNTRY, TOL_SECTOR = '설정!$B$21', '설정!$B$22'

# ══════════════════════════════════════════ 3. 종목정보
ws = wb.create_sheet('종목정보')
head(ws, '종목정보', '가진 종목을 등록하는 곳. 현재가는 앱에서 보고 옮겨 적으세요.',
     {'A': 14, 'B': 18, 'C': 8, 'D': 8, 'E': 16, 'F': 14, 'G': 34})
hrow(ws, 3, ['티커', '종목명', '국가', '통화', '섹터', '현재가(현지통화)', '메모'])
ASSETS = [
    ('005930.KS', '삼성전자', 'KR', 'KRW', '반도체', 85666, '예시 - 지우고 형님 것을 넣으세요'),
    ('373220.KS', 'LG에너지솔루션', 'KR', 'KRW', '2차전지', 372000, '예시'),
    ('AAPL', '애플', 'US', 'USD', 'IT하드웨어', 232.8, '예시'),
    ('AMZN', '아마존', 'US', 'USD', '소비재/클라우드', 178.2, '예시'),
    ('TSLA', '테슬라', 'US', 'USD', '자동차', 250.75, '예시'),
    ('VIC.VN', '빈그룹', 'VN', 'VND', '지주/부동산', 41000, '예시'),
]
for i in range(AST_TOP, AST_END + 1):
    row = ASSETS[i - AST_TOP] if i - AST_TOP < len(ASSETS) else ('',) * 7
    for j, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=val if val != '' else None)
        c.font = NOTE if j == 7 else IN
        c.border = BOX
        if j <= 6: c.fill = IFILL
        if j == 6: c.number_format = PRICE
ws['F3'].comment = Comment('앱 → 현황 화면에서 현재가를 보고 옮겨 적으세요.\n'
                           '비워두면 산 값(매입가)으로 계산합니다.', '찬홍팍')

# ══════════════════════════════════════════ 4. 거래입력
ws = wb.create_sheet('거래입력')
head(ws, '거래입력', '산 날·판 날을 한 줄씩. ★ 반드시 날짜 순으로 적으세요 — '
                     '평단이 순서대로 굴러가기 때문입니다.',
     {'A': 12, 'B': 16, 'C': 13, 'D': 8, 'E': 12, 'F': 12, 'G': 10, 'H': 10, 'I': 16,
      'J': 3, 'K': 11, 'L': 11, 'M': 13, 'N': 13, 'O': 12, 'P': 13})
hrow(ws, 3, ['일자', '종목명', '티커', '구분', '수량', '단가', '수수료', '계좌', '메모'])
hrow(ws, 3, ['이전수량', '누적수량', '이전원가(현지)', '누적원가(현지)',
             '평단(현지)', '실현손익(현지)'], start=11)
for col in 'KLMNOP':
    ws[f'{col}3'].fill = PatternFill('solid', fgColor='667085')
ws['K2'] = '↓ 여기부터는 자동 계산 (건드리지 마세요)'; ws['K2'].font = NOTE

TRADES = [
    ('2024-03-04', '삼성전자', '005930.KS', '매수', 50, 72500, 1090, '연금저축', ''),
    ('2024-05-20', '애플',     'AAPL',      '매수', 20, 189.5,  0.95, '해외주식', ''),
    ('2024-06-11', '아마존',   'AMZN',      '매수', 10, 183.2,  0.92, '해외주식', ''),
    ('2024-07-02', '테슬라',   'TSLA',      '매수', 12, 210.4,  1.05, '해외주식', ''),
    ('2024-09-13', 'LG에너지솔루션', '373220.KS', '매수', 3, 401000, 1800, '연금저축', ''),
    ('2024-11-08', '빈그룹',   'VIC.VN',    '매수', 500, 43200,  108, '해외주식', ''),
    ('2025-01-15', '삼성전자', '005930.KS', '매수', 30, 55900,   840, '연금저축', '추가매수'),
    ('2025-04-22', '테슬라',   'TSLA',      '매도', 5,  268.0,   1.34, '해외주식', '일부 익절'),
]
dv = DataValidation(type='list', formula1='"매수,매도"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'D{TXN_TOP}:D{TXN_END}')

for i in range(TXN_TOP, TXN_END + 1):
    t = TRADES[i - TXN_TOP] if i - TXN_TOP < len(TRADES) else ('',) * 9
    for j, val in enumerate(t, start=1):
        c = ws.cell(row=i, column=j, value=val if val != '' else None)
        c.font = IN; c.border = BOX; c.fill = IFILL
        if j == 1: c.number_format = '@'
        if j == 5: c.number_format = QTY
        if j in (6, 7): c.number_format = PRICE

    p = i - 1
    prevq = '0' if i == TXN_TOP else \
        f'IFERROR(LOOKUP(2,1/($C${TXN_TOP}:$C{p}=$C{i}),$L${TXN_TOP}:$L{p}),0)'
    prevc = '0' if i == TXN_TOP else \
        f'IFERROR(LOOKUP(2,1/($C${TXN_TOP}:$C{p}=$C{i}),$N${TXN_TOP}:$N{p}),0)'
    g = f'IF($C{i}="","",'                                   # 빈 줄은 비워둔다
    # 판 수량은 가진 만큼만 (앱과 같이 초과 매도를 깎는다)
    sold = f'MIN($E{i},$K{i})'
    avg  = f'IF($K{i}=0,0,$M{i}/$K{i})'
    ws[f'K{i}'] = f'={g}{prevq})'
    ws[f'L{i}'] = f'={g}IF($D{i}="매수",$K{i}+$E{i},MAX(0,$K{i}-$E{i})))'
    ws[f'M{i}'] = f'={g}{prevc})'
    ws[f'N{i}'] = (f'={g}IF($D{i}="매수",$M{i}+$E{i}*$F{i}+$G{i},'
                   f'$M{i}-{avg}*{sold}))')
    ws[f'O{i}'] = f'={g}IF($L{i}=0,0,$N{i}/$L{i}))'
    ws[f'P{i}'] = f'={g}IF($D{i}="매수",0,{sold}*$F{i}-{avg}*{sold}-$G{i}))'
    for col, fmt in (('K', QTY), ('L', QTY), ('M', MONEY), ('N', MONEY),
                     ('O', PRICE), ('P', MONEY)):
        c = ws[f'{col}{i}']; c.font = INK; c.border = BOX; c.number_format = fmt
ws.freeze_panes = 'A4'

# ══════════════════════════════════════════ 5. 보유현황
ws = wb.create_sheet('보유현황')
head(ws, '보유현황', '전부 자동 계산입니다. 종목정보에 등록한 순서대로 나옵니다.',
     {'A': 13, 'B': 17, 'C': 7, 'D': 16, 'E': 7, 'F': 11, 'G': 13, 'H': 13, 'I': 10,
      'J': 14, 'K': 14, 'L': 13, 'M': 9, 'N': 8, 'O': 12, 'P': 26})
SUM_CELLS = [('A3', '총 매입금액', 'B3', f'=SUM(J{POS_TOP}:J{POS_END})', MONEY),
             ('C3', '총 평가금액', 'D3', f'=SUM(K{POS_TOP}:K{POS_END})', MONEY),
             ('E3', '평가손익',   'F3', '=D3-B3', MONEY),
             ('G3', '수익률',     'H3', '=IFERROR(F3/B3,"")', PCT),
             ('I3', '실현손익',   'J3', f'=SUM(O{POS_TOP}:O{POS_END})', MONEY)]
for lc, lbl, vc, fml, fmt in SUM_CELLS:
    ws[lc] = lbl; ws[lc].font = H2; ws[lc].fill = SFILL; ws[lc].border = BOX
    ws[vc] = fml; ws[vc].font = INK; ws[vc].fill = SFILL; ws[vc].border = BOX
    ws[vc].number_format = fmt
TOTAL = '보유현황!$D$3'

hrow(ws, 5, ['티커', '종목명', '국가', '섹터', '통화', '보유수량', '평단(현지)',
             '현재가(현지)', '환율', '매입금액(원)', '평가금액(원)', '평가손익(원)',
             '수익률', '비중', '실현손익(원)', '비고'])
for i in range(POS_TOP, POS_END + 1):
    a = AST_TOP + (i - POS_TOP)
    blank = f'IF(종목정보!$A{a}="","",'
    # 엑셀은 빈 칸을 참조하면 "" 가 아니라 0 을 준다. 그대로 넘기면
    # 현재가를 비워둔 종목이 '현재가 0원' 으로 잡혀 평가금액이 0 이 된다.
    # (그래서 한 번 더 빈칸 검사를 씌운다)
    keep = lambda col: f'{blank}IF(종목정보!${col}{a}="","",종목정보!${col}{a}))'
    ws[f'A{i}'] = f'={keep("A")}'
    ws[f'B{i}'] = f'={keep("B")}'
    ws[f'C{i}'] = f'={keep("C")}'
    ws[f'D{i}'] = f'={keep("E")}'
    ws[f'E{i}'] = f'={keep("D")}'
    TX = f'거래입력!$C${TXN_TOP}:$C${TXN_END}'
    ws[f'F{i}'] = (f'={blank}IFERROR(LOOKUP(2,1/({TX}=$A{i}),'
                   f'거래입력!$L${TXN_TOP}:$L${TXN_END}),0))')
    ws[f'G{i}'] = (f'={blank}IFERROR(LOOKUP(2,1/({TX}=$A{i}),'
                   f'거래입력!$O${TXN_TOP}:$O${TXN_END}),0))')
    ws[f'H{i}'] = f'={keep("F")}'
    ws[f'I{i}'] = (f'={blank}IFERROR(INDEX(설정!$B${FX_A}:$B${FX_B},'
                   f'MATCH($E{i},설정!$A${FX_A}:$A${FX_B},0)),""))')
    ws[f'J{i}'] = f'={blank}IFERROR($F{i}*$G{i}*$I{i},0))'
    # 시세가 없으면 산 값으로 친다 (앱과 같은 규칙)
    ws[f'K{i}'] = f'={blank}IFERROR(IF($H{i}="",$J{i},$F{i}*$H{i}*$I{i}),0))'
    ws[f'L{i}'] = f'={blank}IF($H{i}="","",$K{i}-$J{i}))'
    ws[f'M{i}'] = f'={blank}IFERROR($L{i}/$J{i},""))'
    ws[f'N{i}'] = f'={blank}IFERROR($K{i}/{TOTAL},0))'
    # 거래입력의 실현손익은 종목 통화다. 기준통화로 바꿔야 총합이 맞는다.
    ws[f'O{i}'] = (f'={blank}IFERROR(SUMIFS(거래입력!$P${TXN_TOP}:$P${TXN_END},'
                   f'{TX},$A{i})*$I{i},0))')
    ws[f'P{i}'] = (f'={blank}IF($I{i}="","설정 시트에 환율을 넣으세요",'
                   f'IF($F{i}=0,"보유 없음",IF($H{i}="","시세 없음 - 매입가로 계산",""))))')
    for col, fmt, fnt in (('A', '@', LINK), ('B', '@', LINK), ('C', '@', LINK),
                          ('D', '@', LINK), ('E', '@', LINK), ('F', QTY, INK),
                          ('G', PRICE, INK), ('H', PRICE, LINK), ('I', '#,##0.######', LINK),
                          ('J', MONEY, INK), ('K', MONEY, INK), ('L', MONEY, INK),
                          ('M', PCT, INK), ('N', PCT, INK), ('O', MONEY, INK), ('P', '@', INK)):
        c = ws[f'{col}{i}']; c.font = fnt; c.border = BOX; c.number_format = fmt
ws.freeze_panes = 'A6'

# ══════════════════════════════════════════ 6~7. 국가별 / 섹터별
def alloc_sheet(name, src_col, tol_ref, seed, note):
    s = wb.create_sheet(name)
    head(s, name, note,
         {'A': 16, 'B': 16, 'C': 11, 'D': 11, 'E': 11, 'F': 14, 'G': 16, 'H': 34})
    hrow(s, 3, ['항목', '평가금액(원)', '현재 비중', '목표 비중', '차이',
                '판정', '조정 금액(원)', '메모'])
    for i in range(TGT_TOP, TGT_END + 1):
        key = seed[i - TGT_TOP] if i - TGT_TOP < len(seed) else ('', None, '')
        s.cell(row=i, column=1, value=key[0] or None).font = IN
        s.cell(row=i, column=1).fill = IFILL
        t = s.cell(row=i, column=4, value=key[1])
        t.font = IN; t.fill = IFILL; t.number_format = PCT
        s.cell(row=i, column=8, value=key[2] or None).font = NOTE
        g = f'IF($A{i}="","",'
        s[f'B{i}'] = (f'={g}SUMIFS(보유현황!$K${POS_TOP}:$K${POS_END},'
                      f'보유현황!${src_col}${POS_TOP}:${src_col}${POS_END},$A{i}))')
        s[f'C{i}'] = f'={g}IFERROR($B{i}/{TOTAL},0))'
        s[f'D{i}'].border = BOX
        s[f'E{i}'] = f'={g}IF($D{i}="","",$C{i}-$D{i}))'
        s[f'F{i}'] = (f'={g}IF($D{i}="","목표 없음",'
                      f'IF($C{i}>$D{i}+{tol_ref},"비중 많음",'
                      f'IF($C{i}<$D{i}-{tol_ref},"비중 적음","적정"))))')
        s[f'G{i}'] = f'={g}IF($D{i}="","",-$E{i}*{TOTAL}))'
        for col, fmt in (('B', MONEY), ('C', PCT), ('E', PCT), ('F', '@'), ('G', MONEY)):
            c = s[f'{col}{i}']; c.font = INK; c.border = BOX; c.number_format = fmt
        for col in 'AD':
            s[f'{col}{i}'].border = BOX
    r = TGT_END + 1
    s.cell(row=r, column=1, value='합계').font = H2
    for col, fml, fmt in (('B', f'=SUM(B{TGT_TOP}:B{TGT_END})', MONEY),
                          ('C', f'=SUM(C{TGT_TOP}:C{TGT_END})', PCT),
                          ('D', f'=SUM(D{TGT_TOP}:D{TGT_END})', PCT)):
        c = s[f'{col}{r}']; c.value = fml; c.font = H2; c.fill = SFILL
        c.border = BOX; c.number_format = fmt
    s.cell(row=r, column=1).fill = SFILL; s.cell(row=r, column=1).border = BOX
    s.cell(row=r + 2, column=1,
           value='조정 금액: + 는 더 사야 할 금액, − 는 팔아야 할 금액입니다.').font = NOTE
    s.cell(row=r + 3, column=1,
           value='아직 안 가진 항목도 적어두면 "비중 적음" 으로 잡아줍니다. '
                 '(예: 미국 주식이 하나도 없을 때)').font = NOTE
    s.freeze_panes = 'A4'
    return s

alloc_sheet('국가별', 'C', TOL_COUNTRY,
            [('KR', 0.5, '한국'), ('US', 0.5, '미국'), ('VN', None, '베트남 - 목표를 넣으면 판정합니다')],
            '국가별 비중과 목표. 노란 칸(항목·목표 비중)만 고치세요.')
alloc_sheet('섹터별', 'D', TOL_SECTOR,
            [('반도체', 0.3, ''), ('IT하드웨어', 0.2, ''), ('소비재/클라우드', 0.15, ''),
             ('자동차', 0.15, ''), ('2차전지', 0.1, ''), ('지주/부동산', 0.1, '')],
            '섹터별 비중과 목표. 섹터 이름은 종목정보 시트에 적은 것과 똑같이 쓰세요.')

wb.move_sheet('설정', offset=5)
out = str(Path(__file__).resolve().parent.parent / '찬홍팍_주식관리.xlsx')
wb.save(out)
print('saved', out, '| 시트:', wb.sheetnames)
