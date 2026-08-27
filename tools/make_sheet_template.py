# -*- coding: utf-8 -*-
"""올려주신 양식에 수식을 채워 넣는다. 구글시트에 바로 붙여 쓰는 용도.

  - 원본의 칸 배치·병합·열너비를 그대로 두고 수식만 채운다
  - 채워야 하는 칸은 초록 음영
  - 현재가는 구글시트 GOOGLEFINANCE 로 자동, 안 되는 종목은 직접 입력
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

SRC = str(Path(__file__).resolve().parent.parent / 'docs' / '양식_원본.xlsx')
OUT = str(Path(__file__).resolve().parent.parent / '찬홍팍_자산배분_구글시트.xlsx')

FN = 'Arial'
INK  = Font(name=FN, size=10)
IN   = Font(name=FN, size=10, color='0B5394')          # 직접 입력
TH   = Font(name=FN, size=10, bold=True, color='FFFFFF')
H2   = Font(name=FN, size=10, bold=True)
NOTE = Font(name=FN, size=9, color='666666')
GREEN = PatternFill('solid', fgColor='D9EAD3')          # 채워야 하는 칸
HEAD  = PatternFill('solid', fgColor='38761D')
SUB   = PatternFill('solid', fgColor='EFEFEF')
THIN  = Side(style='thin', color='CCCCCC')
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY, PRICE, QTY, PCT = '#,##0', '#,##0.##', '#,##0.####', '0.0%'

ACC_TOP, ACC_END = 2, 54          # 계좌별 보유 (왼쪽 블록)
MST_TOP, MST_END = 6, 30          # 종목 리포트 (오른쪽 블록)
SUM_ROW = MST_END + 2             # 32행: 합계
# 계좌 묶음 (A열 병합과 같은 범위)
GROUPS = [(2, 8), (9, 15), (16, 22), (23, 29), (30, 36), (37, 43), (44, 50),
          (51, 51), (52, 53), (54, 54)]
# 종목 이름에서 바로 읽히는 자산군만 채운다 (원본 L1:T1 머리글과 같은 말)
CLASS = {
    'ACE 미국나스닥100': '미국', 'TIGER  미국S&P500': '미국',
    'SOL 미국배당다우존스': '미국', 'TIGER  미국배당다우존스': '미국',
    'KODEX 미국나스닥100': '미국', 'KODEX 미국S&P500': '미국',
    '1Q미국S&P500미국채혼합50액티브': '미국', '매그니피센트7 MAGS': '미국',
    'Schwab SCHD': '미국', '버크셔 해서웨이B BRK.B': '미국',
    'KODEX 200TR': '한국', 'TIGER 일본니케이225': '일본',
    'KODEX 차이나CSI300': '중국', 'KODEX 인도Nifty50': '인도',
    '금 99.99K': '금', '비트코인': '비트코인', '이더리움': '이더리움', '현금': '현금',
}

wb = load_workbook(SRC)
ws = wb['Sheet1']
ws.title = '자산배분'

# 원본에 왼쪽은 '99.99K', 오른쪽 종목표는 '금 99.99K' 로 적혀 있어 서로 못 찾는다.
# 이름이 다르면 그 보유가 통째로 계산에서 빠지므로 오른쪽 표 이름에 맞춘다.
if ws['B51'].value == '99.99K':
    ws['B51'] = '금 99.99K'

def put(ref, value, font=INK, fill=None, fmt=None, align='right'):
    c = ws[ref]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    c.border = BOX
    c.alignment = Alignment(horizontal=align, vertical='center')
    return c

def header(ref, text):
    c = ws[ref]; c.value = text; c.font = TH; c.fill = HEAD; c.border = BOX
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ═══════════════ 머리글
for ref, t in [('A1', '계좌구분'), ('B1', '종목명'), ('C1', '현재가'), ('D1', '매입가'),
               ('E1', '수량'), ('F1', '평가금액'), ('G1', '계')]:
    header(ref, t)
for ref, t in [('I5', '종목'), ('J5', '종목코드'), ('K5', '현재가'), ('L5', '매입가'),
               ('M5', '대비'), ('N5', '등락률'), ('O5', '갱신시간'), ('P5', '수량'),
               ('Q5', '평가금액'), ('R5', '현재비중'), ('S5', '목표비중'),
               ('T5', '목표금액'), ('U5', '매매금액'), ('V5', '매매량'),
               ('W5', '자산군'), ('X5', '수동 현재가')]:
    header(ref, t)
for ref in ('L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1'):
    c = ws[ref]; c.font = H2; c.fill = SUB; c.border = BOX
    c.alignment = Alignment(horizontal='center')
header('U1', '합계')
ws['I1'].font = H2; ws['I1'].alignment = Alignment(horizontal='center')
for ref in ('K2', 'K3'):
    ws[ref].font = H2; ws[ref].alignment = Alignment(horizontal='right')

MST = f'$I${MST_TOP}:$I${MST_END}'
TOT = f'$Q${SUM_ROW}'

# ═══════════════ 왼쪽: 계좌별 보유
for r in range(ACC_TOP, ACC_END + 1):
    b = ws[f'B{r}']
    b.font = IN; b.fill = GREEN; b.border = BOX
    b.alignment = Alignment(horizontal='left', vertical='center')
    g = f'IF($B{r}="","",'
    put(f'C{r}', f'={g}IFERROR(INDEX($K${MST_TOP}:$K${MST_END},'
                 f'MATCH($B{r},{MST},0)),"종목표에 없음"))', fmt=PRICE)
    put(f'D{r}', None, font=IN, fill=GREEN, fmt=PRICE)
    put(f'E{r}', None, font=IN, fill=GREEN, fmt=QTY)
    put(f'F{r}', f'={g}IFERROR($C{r}*$E{r},""))', fmt=MONEY)
ws['B1'].comment = Comment(
    '오른쪽 [종목] 표에 있는 이름과 똑같이 적어야 현재가가 붙습니다.\n'
    '이름이 다르면 현재가 칸에 "종목표에 없음" 이라고 뜹니다.', '찬홍팍')

# 계좌별 합계 (A열 병합과 같은 범위로 묶는다)
for a, z in GROUPS:
    if z > a: ws.merge_cells(f'G{a}:G{z}')
    put(f'G{a}', f'=SUM($F{a}:$F{z})', font=H2, fmt=MONEY, align='center')
    ws[f'G{a}'].alignment = Alignment(horizontal='center', vertical='center')
    for r in range(a, z + 1):
        ws[f'A{r}'].border = BOX; ws[f'G{r}'].border = BOX

# ═══════════════ 오른쪽 위: 투자 합계 / 목표·현재 비율
ws.merge_cells('I2:J2'); ws.merge_cells('I3:J3')
put('I2', f'=SUM($F${ACC_TOP}:$F${ACC_END})', font=H2, fmt=MONEY, align='center')
orphan = (f'SUMPRODUCT(($B${ACC_TOP}:$B${ACC_END}<>"")'
          f'*(COUNTIF({MST},$B${ACC_TOP}:$B${ACC_END})=0))')
put('I3', f'=IF({orphan}>0,"왼쪽 종목명 "&{orphan}&"개가 오른쪽 표에 없습니다",'
          f'IF(ROUND(SUM($F${ACC_TOP}:$F${ACC_END}),0)=ROUND({TOT},0),'
          f'"계좌합계 = 종목합계 OK","계좌합계와 종목합계가 다릅니다"))',
    font=NOTE, align='center')
for col in 'LMNOPQRST':
    put(f'{col}2', ws[f'{col}2'].value, font=IN, fill=GREEN, fmt=PCT)
    put(f'{col}3', f'=IFERROR(SUMIF($W${MST_TOP}:$W${MST_END},{col}$1,'
                   f'$Q${MST_TOP}:$Q${MST_END})/{TOT},"")', fmt=PCT)
put('U2', '=SUM($L$2:$T$2)', font=H2, fmt=PCT)
put('U3', '=SUM($L$3:$T$3)', font=H2, fmt=PCT)
put('V2', '=IF(ROUND($U$2,4)=1,"","목표비율 합이 100% 가 아닙니다")',
    font=Font(name=FN, size=9, color='CC0000'), align='left')

# ═══════════════ 오른쪽: 종목 리포트
for r in range(MST_TOP, MST_END + 1):
    name = ws[f'I{r}'].value
    i = ws[f'I{r}']; i.font = IN; i.fill = GREEN; i.border = BOX
    i.alignment = Alignment(horizontal='left', vertical='center')
    put(f'J{r}', None, font=IN, fill=GREEN, align='left')
    put(f'W{r}', CLASS.get(name), font=IN, fill=GREEN, align='center')
    put(f'X{r}', None, font=IN, fill=GREEN, fmt=PRICE)

    g = f'IF($I{r}="","",'
    gf = f'IF($J{r}="","",'
    # 종목코드가 있으면 구글파이낸스, 없거나 안 되면 '수동 현재가' 를 쓴다
    put(f'K{r}', f'={g}IF($J{r}="",$X{r},IFERROR(GOOGLEFINANCE($J{r},"price"),$X{r})))',
        fmt=PRICE)
    put(f'L{r}', f'={g}IFERROR(SUMPRODUCT(($B${ACC_TOP}:$B${ACC_END}=$I{r})'
                 f'*$D${ACC_TOP}:$D${ACC_END}*$E${ACC_TOP}:$E${ACC_END})/$P{r},""))', fmt=PRICE)
    put(f'M{r}', f'={gf}IFERROR(GOOGLEFINANCE($J{r},"price")'
                 f'-GOOGLEFINANCE($J{r},"closeyest"),""))', fmt=PRICE)
    put(f'N{r}', f'={gf}IFERROR(GOOGLEFINANCE($J{r},"changepct")/100,""))', fmt=PCT)
    put(f'O{r}', f'={gf}IFERROR(GOOGLEFINANCE($J{r},"tradetime"),""))',
        fmt='mm-dd hh:mm')
    put(f'P{r}', f'={g}SUMIF($B${ACC_TOP}:$B${ACC_END},$I{r},'
                 f'$E${ACC_TOP}:$E${ACC_END}))', fmt=QTY)
    put(f'Q{r}', f'={g}IFERROR($K{r}*$P{r},""))', fmt=MONEY)
    put(f'R{r}', f'={g}IFERROR($Q{r}/{TOT},""))', fmt=PCT)
    # 자산군 목표를, 그 자산군 안에서는 지금 금액 비율대로 나눠 준다.
    # 그 자산군이 아직 비어 있으면 종목 수대로 똑같이 나눈다.
    cls_sum = f'SUMIF($W${MST_TOP}:$W${MST_END},$W{r},$Q${MST_TOP}:$Q${MST_END})'
    cls_cnt = f'COUNTIF($W${MST_TOP}:$W${MST_END},$W{r})'
    tgt = f'INDEX($L$2:$T$2,MATCH($W{r},$L$1:$T$1,0))'
    put(f'S{r}', f'={g}IF($W{r}="","",IFERROR({tgt}*IF({cls_sum}=0,'
                 f'1/{cls_cnt},$Q{r}/{cls_sum}),"")))', fmt=PCT)
    put(f'T{r}', f'={g}IFERROR($S{r}*{TOT},""))', fmt=MONEY)
    put(f'U{r}', f'={g}IFERROR($T{r}-$Q{r},""))', fmt=MONEY)
    put(f'V{r}', f'={g}IFERROR($U{r}/$K{r},""))', fmt=QTY)

put(f'I{SUM_ROW}', '합계', font=H2, align='center')
ws[f'I{SUM_ROW}'].fill = SUB
for col, fmt in (('P', QTY), ('Q', MONEY), ('R', PCT), ('S', PCT),
                 ('T', MONEY), ('U', MONEY)):
    c = put(f'{col}{SUM_ROW}', f'=SUM({col}${MST_TOP}:{col}${MST_END})',
            font=H2, fmt=fmt)
    c.fill = SUB
for col in 'JKLMNOV':
    ws[f'{col}{SUM_ROW}'].fill = SUB; ws[f'{col}{SUM_ROW}'].border = BOX

# 자산군은 위 머리글에서 고르게 (구글시트에서 드롭다운이 된다)
dv = DataValidation(type='list', formula1='=$L$1:$T$1', allow_blank=True)
ws.add_data_validation(dv); dv.add(f'W{MST_TOP}:W{MST_END}')
dv2 = DataValidation(type='list', formula1=f'={MST}', allow_blank=True)
ws.add_data_validation(dv2); dv2.add(f'B{ACC_TOP}:B{ACC_END}')

for col, w in {'C': 11, 'D': 11, 'E': 10, 'F': 13, 'G': 13, 'H': 2, 'J': 13, 'K': 11,
               'L': 11, 'M': 9, 'N': 9, 'O': 12, 'P': 10, 'Q': 13, 'R': 10, 'S': 10,
               'T': 13, 'U': 13, 'V': 11, 'W': 11, 'X': 12}.items():
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'B2'
ws.sheet_view.showGridLines = False

# ═══════════════ 사용법
g = wb.create_sheet('사용법')
g.sheet_view.showGridLines = False
g.column_dimensions['A'].width = 20
g.column_dimensions['B'].width = 104
g['A1'] = '이 표 쓰는 법'; g['A1'].font = Font(name=FN, size=14, bold=True)
LINES = [
    ('', ''),
    ('■ 초록 칸만 채우세요', '나머지는 전부 수식입니다. 지우면 계산이 깨집니다.'),
    ('', ''),
    ('  왼쪽 (A~G)', '계좌별로 뭘 얼마에 몇 주 샀는지 적는 곳입니다.'),
    ('  · 종목명(B)', '오른쪽 [종목] 표의 이름과 똑같이. 칸을 누르면 목록에서 고를 수 있습니다.'),
    ('  · 매입가(D)', '그 계좌에서 산 평균 단가.'),
    ('  · 수량(E)', '보유 주수.'),
    ('  → 현재가·평가금액·계', '자동으로 나옵니다.'),
    ('', ''),
    ('  오른쪽 (I~X)', '종목 하나당 한 줄인 통합표입니다.'),
    ('  · 종목(I)', '종목 이름.'),
    ('  · 종목코드(J)', '구글파이낸스 코드. 넣으면 현재가·대비·등락률·갱신시간이 자동으로 붙습니다.'),
    ('', '    국내  KRX:069500      (여섯 자리 종목코드)'),
    ('', '    미국  NASDAQ:AAPL / NYSEARCA:SCHD / NYSE:BRK.B'),
    ('', '    코드는 구글에서 종목명을 검색하면 나옵니다. 확인하고 넣으세요.'),
    ('  · 자산군(W)', '미국/일본/중국/한국/인도/금/비트코인/이더리움/현금 중에서 고릅니다.'),
    ('', '    이게 있어야 목표비율과 맞춰볼 수 있습니다. 칸을 누르면 목록이 나옵니다.'),
    ('  · 수동 현재가(X)', '종목코드가 없거나 구글파이낸스가 못 찾는 종목(금 현물 등)의 가격.'),
    ('', '    종목코드가 있으면 이 칸은 무시됩니다.'),
    ('', ''),
    ('  위쪽 (L2~T2)', '자산군별 목표비율. 여기만 고치면 아래 계산이 전부 따라옵니다.'),
    ('', ''),
    ('■ 자동으로 나오는 것', ''),
    ('  현재비중(R)', '그 종목이 전체에서 차지하는 비율.'),
    ('  목표비중(S)', '그 자산군 목표를, 자산군 안에서는 지금 금액 비율대로 나눈 값입니다.'),
    ('', '    (그 자산군이 아직 비어 있으면 종목 수대로 똑같이 나눕니다)'),
    ('  목표금액(T)', '목표비중 × 전체 금액.'),
    ('  매매금액(U)', '목표금액 − 지금 평가금액.  + 면 그만큼 더 사고, − 면 그만큼 팝니다.'),
    ('  매매량(V)', '매매금액 ÷ 현재가. 몇 주 사고팔면 되는지.'),
    ('  I3 칸', '왼쪽 계좌합계와 오른쪽 종목합계가 맞는지 스스로 검사합니다.'),
    ('', '    "다릅니다" 가 뜨면 왼쪽 종목명 중에 오른쪽 표에 없는 게 있다는 뜻입니다.'),
    ('', ''),
    ('■ 구글시트로 옮기기', '드라이브에 이 파일을 올리고 → 구글 스프레드시트로 열기.'),
    ('', '현재가가 붙는 GOOGLEFINANCE 는 구글시트에서만 돕니다.'),
    ('', '엑셀에서 열면 그 칸들만 #NAME? 으로 보입니다 (구글시트에 올리면 정상).'),
    ('', ''),
    ('■ 확인해 주세요', '지금 목표비율(L2~T2) 합이 150% 입니다. 100% 가 되게 고쳐 주세요.'),
    ('', 'U2 칸에 합이 나오고, 100% 가 아니면 옆에 빨간 글씨로 알려줍니다.'),
    ('', ''),
    ('■ 넣을 수 있는 양', '계좌별 보유 53줄 / 종목 25개. 모자라면 마지막 줄을 복사해 늘리세요.'),
]
r = 3
for a, b in LINES:
    g.cell(row=r, column=1, value=a).font = Font(name=FN, size=10, bold=a.startswith('■'))
    c = g.cell(row=r, column=2, value=b); c.font = Font(name=FN, size=10)
    c.alignment = Alignment(vertical='top')
    r += 1
g.cell(row=r + 1, column=1, value='■■■').font = Font(name=FN, size=10)
g.cell(row=r + 1, column=1).fill = GREEN
g.cell(row=r + 1, column=2, value='이 색 = 채우는 칸').font = Font(name=FN, size=10)

Path(OUT).parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print('saved', OUT)
